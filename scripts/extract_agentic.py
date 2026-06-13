#!/usr/bin/env python3
"""Extract the curated Agentic-AI tools spreadsheet into data/agentic-tools.json.

Usage:
    uv run --with openpyxl python scripts/extract_agentic.py path/to/agentic.xlsx

The catalog generator (scripts/build-catalog.ts) reads the JSON as a second
source: each row becomes an "AI & Agents" catalog entry, screened by GitHub
owner against the exclusion policy (so OpenAI/Meta/xAI-owned tools are dropped).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl required. Run: uv run --with openpyxl python scripts/extract_agentic.py <xlsx>")


def col_index(header, *prefixes):
    for i, h in enumerate(header):
        hl = str(h or "").strip().lower()
        for p in prefixes:
            if hl.startswith(p.lower()):
                return i
    return None


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit("Usage: extract_agentic.py path/to/agentic.xlsx")
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"No such file: {src}")

    wb = load_workbook(src, read_only=True, data_only=True)
    if "All Repositories" not in wb.sheetnames:
        sys.exit(f"No 'All Repositories' sheet. Sheets: {wb.sheetnames}")
    ws = wb["All Repositories"]
    rows = list(ws.iter_rows(values_only=True))

    # Header is the first row that contains a 'Repository' cell.
    header_idx = next((i for i, r in enumerate(rows) if r and any(str(c or "").strip() == "Repository" for c in r)), None)
    if header_idx is None:
        sys.exit("Could not locate the header row (no 'Repository' column).")
    header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]

    ci = {
        "name": col_index(header, "Repository"),
        "category": col_index(header, "Category"),
        "subcategory": col_index(header, "Sub-Category", "Sub Category"),
        "lang": col_index(header, "Language"),
        "license": col_index(header, "License"),
        "backer": col_index(header, "Primary Backer"),
        "url": col_index(header, "Source URL", "URL"),
        "desc": col_index(header, "Description"),
    }
    if ci["name"] is None or ci["url"] is None:
        sys.exit(f"Missing required columns. Header: {header}")

    def cell(r, key):
        i = ci[key]
        if i is None or i >= len(r) or r[i] is None:
            return ""
        return str(r[i]).strip()

    out = []
    for r in rows[header_idx + 1:]:
        if not r or not any(c is not None and str(c).strip() for c in r):
            continue
        name = cell(r, "name")
        if not name:
            continue
        out.append({
            "name": name,
            "category_label": cell(r, "category"),
            "subcategory": cell(r, "subcategory"),
            "language": cell(r, "lang"),
            "license_hint": cell(r, "license"),
            "backer": cell(r, "backer"),
            "source_url": cell(r, "url"),
            "description": cell(r, "desc"),
        })

    Path("data").mkdir(parents=True, exist_ok=True)
    Path("data/agentic-tools.json").write_text(json.dumps(out, indent=2, ensure_ascii=False))
    print(f"data/agentic-tools.json: {len(out)} tools")


if __name__ == "__main__":
    main()
