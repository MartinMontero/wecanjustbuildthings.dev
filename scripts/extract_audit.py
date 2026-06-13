#!/usr/bin/env python3
"""Extract the AOS dependency audit spreadsheet into CSVs the catalog
generator consumes.

Usage:
    uv run --with openpyxl python scripts/extract_audit.py path/to/audit.xlsx

Produces:
    data/aos-dependency-audit.csv   (the ALL DEPENDENCIES sheet)
    data/aos-repos.csv              (the PER-REPO DEEP DIVE sheet, if present)

The script is deliberately tolerant: it warns on a row-count or header mismatch
rather than aborting, so a slightly different export still yields usable data.
Re-running `npm run data:fetch` then turns these rows into verified entries.
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required. Run via: uv run --with openpyxl python scripts/extract_audit.py <xlsx>")

EXPECTED = ["Dependency", "Ecosystem", "Category", "Repos Using", "What It Does", "Used In"]


def main() -> None:
    if len(sys.argv) < 2:
        sys.exit("Usage: extract_audit.py path/to/audit.xlsx")
    src = Path(sys.argv[1])
    if not src.exists():
        sys.exit(f"No such file: {src}")

    out = Path("data")
    out.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(filename=src, read_only=True, data_only=True)

    deps_sheet = _find_sheet(wb, ["ALL DEPENDENCIES", "Dependencies", "All Dependencies"])
    if deps_sheet is None:
        sys.exit(f"Could not find a dependencies sheet. Sheets present: {wb.sheetnames}")

    rows = list(deps_sheet.iter_rows(values_only=True))
    if not rows:
        sys.exit("Dependencies sheet is empty.")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    if header[: len(EXPECTED)] != EXPECTED:
        print(f"WARNING: header mismatch.\n  expected {EXPECTED}\n  got      {header[:len(EXPECTED)]}", file=sys.stderr)

    written = 0
    with open(out / "aos-dependency-audit.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(EXPECTED)
        for r in rows[1:]:
            if not r or r[0] is None or not str(r[0]).strip():
                continue  # skip blanks / section headers
            w.writerow(["" if c is None else str(c).strip() for c in r[: len(EXPECTED)]])
            written += 1
    print(f"data/aos-dependency-audit.csv: {written} rows")
    if written != 1161:
        print(f"NOTE: expected 1161 rows, wrote {written}. Verify the source export.", file=sys.stderr)

    repos_sheet = _find_sheet(wb, ["PER-REPO DEEP DIVE", "Per-Repo Deep Dive", "Repos"])
    if repos_sheet is not None:
        with open(out / "aos-repos.csv", "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for r in repos_sheet.iter_rows(values_only=True):
                if not r or all(c is None for c in r):
                    continue
                w.writerow(["" if c is None else str(c).strip() for c in r])
        print("data/aos-repos.csv written")


def _find_sheet(wb, candidates):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    return None


if __name__ == "__main__":
    main()
